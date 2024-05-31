import openpyxl
from openpyxl.utils import get_column_letter

def create_excel_history(
        nama, 
        umur, 
        gender,
        status,
        pola_makan, 
        perokok, 
        alkohol, 
        obat_terlarang,
        bmi, 
        ahh,
        sisa_umur
    ):
    try:
        workbook = openpyxl.load_workbook("Kalkulator-Ekspektasi-Hidup-Manusia/life_expectancy_history.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Life Expectancy History"
        headers = [
            "Nama", 
            "Umur", 
            "Gender", 
            "Status", 
            "Pola Makan", 
            "Perokok", 
            "Alkohol", 
            "Obat Terlarang", 
            "BMI", 
            "Angka Harapan Hidup", 
            "Sisa Umur"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col).value = header

    data = [
            [
            nama, 
            umur,
            gender,
            status,
            pola_makan, 
            perokok, 
            alkohol, 
            obat_terlarang, 
            bmi, 
            ahh,
            sisa_umur,
        ],

    ]

    headers = [
            "Nama", 
            "Umur", 
            "Gender", 
            "Status", 
            "Pola Makan", 
            "Perokok", 
            "Alkohol", 
            "Obat Terlarang", 
            "BMI", 
            "Angka Harapan Hidup", 
            "Sisa Umur"
        ]

    last_row = sheet.max_row + 1

    for row, row_data in enumerate(data, start=last_row):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=row, column=col).value = value

    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].auto_size = True

    workbook.save("Kalkulator-Ekspektasi-Hidup-Manusia/life_expectancy_history.xlsx")

def display_history(name=None):
    workbook = openpyxl.load_workbook("Kalkulator-Ekspektasi-Hidup-Manusia/life_expectancy_history.xlsx")
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    if name:
        data = [row for row in data if row[0] == name]
    
    if not data:
        print("Data tidak ditemukan.")
        return

    for row in data:
        data = [
            row [0],
            row [9],
            row [10],
        ]
    return data

def read_nama():
    workbook = openpyxl.load_workbook("Kalkulator-Ekspektasi-Hidup-Manusia/life_expectancy_history.xlsx")
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append((row[0], row[2], row[3]))
    return data

def delete_data(nama):
    workbook = openpyxl.load_workbook("Kalkulator-Ekspektasi-Hidup-Manusia/life_expectancy_history.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == nama:
            sheet.delete_rows(row[0].row, 1)
            break

    workbook.save("Kalkulator-Ekspektasi-Hidup-Manusia/life_expectancy_history.xlsx")